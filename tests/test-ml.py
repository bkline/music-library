#!/usr/bin/env python3

"""Test the Music Library Catalog software.

You will need to set up a Python (3.10 or higher) environment in which
the third-party packages openpyxl (3.1.2+), requests (2.31.0+), and
selenium (4.0.0+) are installed. A Python virtual environment is the
most effective way to achieve this.

You will also need to install Chrome for Testing (120.0+) from
https://googlechromelabs.github.io/chrome-for-testing/.
The most efficient way to run the tests involves launching Chrome
for Testing before invoking the test script.

Finally, you will need to supply the credentials for an account that
has write access to the Music Library Catalog.

With a perfect test suite, any of the tests could be run independently.
However, some of the other tests assume that the first test has been run,
creating a test item.
"""

from argparse import ArgumentParser
from base64 import b64decode
from datetime import datetime
from functools import cached_property
from io import BytesIO
from logging import getLogger, Formatter, FileHandler
from re import compile as re_compile
from sys import argv
from time import sleep
from unittest import TestCase, main
from openpyxl import load_workbook
from requests import Session
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


class Tests(TestCase):
    """Suite of automated tests for the Music Library Catalog"""

    DEFAULT_WAIT = 20
    BASE = "https://ml.rksystems.com"
    LOG_FORMAT = "%(asctime)s [%(levelname)s] %(message)s"
    LOG_PATH = "test-library-catalog.log"
    LOGGER = getLogger("catalog-test")
    HANDLER = FileHandler(LOG_PATH, encoding="utf-8")
    HANDLER.setFormatter(Formatter(LOG_FORMAT))
    LOGGER.setLevel("INFO")
    LOGGER.addHandler(HANDLER)
    STARTED = datetime.now()
    VERBOSE = False
    LOADING = re_compile(r"<(div|p)>(Loading|Waiting) ?\.\.\.</\1>")
    LOOKUPS = [
        ("accompaniment", "Accompaniments", "Test Accompaniment Value"),
        ("arrangement", "Arrangements", "Test Arrangement Value"),
        (
            "handbell-ensemble",
            "HandBell Ensembles",
            "Test Handbell Ensemble Value",
        ),
        ("key", "Keys", "Test Key Value"),
        ("keyword", "Keywords", "Test Keyword Value"),
        ("owner", "Owners", "Test Owner Value"),
        ("season", "Seasons", "Test Season Value"),
        ("skill", "Skills", "Test Skill Value"),
        ("tag", "Tags", "Test Tag Name"),
        ("tag-group", "Tag Groups", "Test Tag Group"),
        ("person", "People", "Person, Test"),
        ("company", "Companies", "Test Company Name"),
    ]

    def setUp(self):
        """This gets run at the start of every test."""

        self.started = datetime.now()
        options = webdriver.ChromeOptions()
        if not self.VERBOSE:
            exclude = ["enable-logging"]
            options.add_experimental_option("excludeSwitches", exclude)
        self.driver = webdriver.Chrome(options=options)
        self.driver.implicitly_wait(self.DEFAULT_WAIT)
        self.login()

    def tearDown(self):
        """This gets run at the end of every test."""

        sleep(2)
        self.scroll_to_top()
        button = self.find("logout-button", By.ID)
        self.assertIsNotNone(button)
        button.click()
        self.save_pdf("logout.pdf")
        self.assert_page_has("Thanks for spending quality time")
        self.assert_page_has("Log In")
        self.check_console_errors()
        self.driver.close()
        elapsed = datetime.now() - self.started
        method_name = self.id().split(".")[2]
        self.logger.info("%s elapsed for %s", elapsed, method_name)

    @cached_property
    def logger(self):
        """Record what happens during the tests."""
        return self.LOGGER

    @cached_property
    def open_tabs(self):
        """The browser tabs which have been opened by the current test."""
        return set()

    @cached_property
    def url(self):
        """Base URL for the tests."""
        return f"{self.BASE}/library"

    def assert_page_has(self, expected):
        """Assert that the source text for the page contains a specific value.

        Required positional argument:
          expected - string we should find
        """

        self.assertIn(str(expected), self.get_page_source())

    def assert_page_not_has(self, expected):
        """The reverse of the "page_has" assertion.

        Required positional argument:
          expected - string we should not find
        """

        self.assertNotIn(expected, self.get_page_source())

    def check_console_errors(self):
        """Look for errors in the browser's console."""

        logs = self.driver.get_log("browser")
        for log in logs:
            if log["level"] == "SEVERE":
                self.logger.error(log)
            else:
                self.logger.info(log)

    def find(self, selector, method=By.CSS_SELECTOR, all=False):
        """Find matching element(s).

        Required positional argument:
          selector - string identifying which elements we want

        Optional keyword arguments:
          method - default is By.CSS_SELECTOR
          all - set to True to return all matching elements
        """

        if all:
            return self.driver.find_elements(method, selector)
        return self.driver.find_element(method, selector)

    def get_page_source(self):
        """Wait for the page to be completely loaded."""

        attempts = 3
        sleep(2)
        while attempts > 0:
            source = self.driver.page_source
            if "</body>" in source and "</html>" in source:
                if not self.LOADING.search(source):
                    return source
                else:
                    message = "...waiting"
                    args = []
                    self.logger.warning("...waiting")
            else:
                args.append(source)
                message = "body missing in %s"
            attempts -= 1
            args.append(attempts)
            message += "; %d tries left"
            self.logger.warning(message, *args)
            if attempts:
                sleep(3 - attempts)
        raise Exception("body element never found")

    def login(self):
        """Create a new login session."""

        self.navigate_to(self.url)
        self.set_field_value("username", self.USERNAME)
        self.set_field_value("password", self.PASSWORD)
        button = self.find("//button[contains(text(), 'Login')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()

    def navigate_to(self, path, **params):
        """Open the page for a given CGI script.

        Required positional argument:
          path - string url's resource path

        Optional keyword arguments:
          params - CGI named parameters needed for the test

        Return:
          string for the handle of the browser tab for the page
        """

        self.logger.debug("navigating to %s", path)
        self.driver.get(path)
        handle = self.driver.current_window_handle
        self.open_tabs.add(handle)
        sleep(1)
        return handle

    def run_report(self, expected="Test Item Title"):
        """Run a report and verify that it found our test record."""

        sleep(1)
        path = "//button[contains(text(), 'Report')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(2)
        original = self.driver.current_window_handle
        button.click()
        sleep(1)
        if expected is not None:
            handles = self.driver.window_handles
            for handle in handles:
                if handle != original:
                    report = handle
                    self.driver.switch_to.window(report)
                    break
            self.assert_page_has(expected)

    def set_field_value(self, field_id, value):
        """Replace a form field's value.

        Required positional arguments:
          field_id - string for the field's unique ID
          value - the new value to be associated with the field
        """

        field = self.driver.find_element(By.ID, field_id)
        if not field:
            self.logger.warning("can't find field %s", field_id)
        else:
            field.clear()
            field.send_keys(str(value))

    def save_pdf(self, filename="page.pdf"):
        """Create a PDF file of the current page.

        Optional keyword argument:
          filename - defaults to 'page.pdf'
        """

        with open(filename, "wb") as fp:
            fp.write(b64decode(self.driver.print_page()))

    def scroll_to_bottom(self):
        """Force the browser to the bottom of the page."""

        get_height = "return document.body.scrollHeight"
        scroll = "window.scrollTo(0, document.body.scrollHeight);"
        last_height = self.driver.execute_script(get_height)
        limit = 5
        while limit > 0:
            self.driver.execute_script(scroll)
            sleep(2)
            new_height = self.driver.execute_script(get_height)
            if new_height == last_height:
                break
            limit -= 1
            last_height = new_height

    def scroll_to_top(self):
        """Force the browser to the top of the page."""

        get_position = "return window.pageYOffset"
        scroll = "window.scrollTo(0, 0);"
        last_position = self.driver.execute_script(get_position)
        limit = 5
        while limit > 0:
            self.driver.execute_script(scroll)
            sleep(2)
            new_position = self.driver.execute_script(get_position)
            if new_position == last_position:
                break
            limit -= 1
            last_position = new_position

    def test_01_create_item(self):
        """Add a new item to the catalog."""

        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.assert_page_has("Search for an existing record to edit")
        button = self.find("//button[contains(text(), 'Create')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        self.assert_page_has("Adding New Record")
        path = "//legend[contains(text(), 'Identification')]"
        legend = self.find(path, By.XPATH)
        self.assertIsNotNone(legend)
        legend.click()
        self.set_field_value("ItemTitle", "Test Item Title")
        self.set_field_value("OtherTitle", "Test Item Other Title")
        self.set_field_value("Comments", "Test Item Comments")
        self.driver.execute_script("arguments[0].scrollIntoView(true)", legend)
        sleep(1)
        legend.click()
        sleep(1)
        path = "//legend[contains(text(), 'Publication')]"
        legend = self.find(path, By.XPATH)
        self.assertIsNotNone(legend)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", legend)
        legend.click()
        sleep(2)
        self.set_field_value("Copyright", "2025 All Rights Reserved")
        self.set_field_value("StockNumber", "Test Item Stock Number")
        legend.click()
        sleep(1)
        path = "//legend[contains(text(), 'Musical Information')]"
        legend = self.find(path, By.XPATH)
        self.assertIsNotNone(legend)
        legend.click()
        sleep(1)
        self.set_field_value("TimeSignature", "3/4")
        self.set_field_value("Duration", "5.25")
        self.set_field_value("ProgramNotes", "Test Item Program Notes")
        legend.click()
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(2)
        button.click()
        self.assert_page_has("Successfully saved Test Item Title")

    def test_02_accompaniment_value(self):
        """Add a new value to the accompaniment lookup table."""

        button = self.find("//a[contains(text(), 'Lookup Tables')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        path = "a[href='/library/lookup/accompaniment']"
        link = self.find(path, By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.assert_page_has("Accompaniments")
        self.assert_page_has("Select a value to edit")
        button = self.find("//button[contains(text(), 'Create')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        self.assert_page_has("Editing Lookup Value")
        self.set_field_value("LookupValue", "Test Accompaniment Value")
        self.set_field_value("SortPosition", "-1000")
        self.set_field_value("Comments", "Yada yada yada...")
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.set_field_value("title", "Test Item Title")
        button = self.find("//button[contains(text(), 'Search')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        button = self.find("button[title='Edit item']", By.CSS_SELECTOR)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        path = "//legend[contains(text(), 'Musical Information')]"
        legend = self.find(path, By.XPATH)
        self.assertIsNotNone(legend)
        legend.click()
        sleep(1)
        field = self.find("AccompanimentID", By.NAME)
        field.send_keys("Test Accompaniment Value")
        sleep(1)
        xpath = "//option[contains(text(), 'Test Accompaniment Value')]"
        option = self.find(xpath, By.XPATH)
        option.click()
        legend.click()
        sleep(1)
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(2)
        button.click()

    def test_03_arrangement_value(self):
        """Add a new value to the arrangement lookup table."""

        button = self.find("//a[contains(text(), 'Lookup Tables')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        path = "a[href='/library/lookup/arrangement']"
        link = self.find(path, By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.assert_page_has("Arrangements")
        self.assert_page_has("Select a value to edit")
        button = self.find("//button[contains(text(), 'Create')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        self.assert_page_has("Editing Lookup Value")
        self.set_field_value("LookupValue", "Test Arrangement Value")
        self.set_field_value("SortPosition", "-1000")
        self.set_field_value("Comments", "Yada yada yada...")
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.set_field_value("title", "Test Item Title")
        button = self.find("//button[contains(text(), 'Search')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        button = self.find("button[title='Edit item']", By.CSS_SELECTOR)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        path = "//legend[contains(text(), 'Musical Information')]"
        legend = self.find(path, By.XPATH)
        self.assertIsNotNone(legend)
        legend.click()
        sleep(1)
        field = self.find("ArrangementID", By.NAME)
        field.send_keys("Test Arrangement Value")
        sleep(1)
        xpath = "//option[contains(text(), 'Test Arrangement Value')]"
        option = self.find(xpath, By.XPATH)
        option.click()
        legend.click()
        sleep(1)
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(2)
        button.click()

    def test_04_company_value(self):
        """Add a new value to the company lookup table."""

        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.set_field_value("title", "Test Item Title")
        button = self.find("//button[contains(text(), 'Search')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        button = self.find("button[title='Edit item']", By.CSS_SELECTOR)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        path = "//button[contains(text(), 'Add Company')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(2)
        self.scroll_to_bottom()
        button.click()
        self.assert_page_has("Create New Company Record")
        self.set_field_value("CompanyName", "Test Company Name")
        self.set_field_value("WebSite", "https://www.example.com")
        self.set_field_value("Address", "23 Skidoo Lane")
        self.set_field_value("City", "Kalamazoo")
        self.set_field_value("State", "MI")
        self.set_field_value("ZIPCode", "49008")
        self.set_field_value("Country", "USA")
        self.set_field_value("Phone", "+1 269-555-1212")
        self.set_field_value("PhoneNotes", "Don't call after 9pm")
        self.set_field_value("Email", "klem@kadiddlehopper.org")
        field = self.find(".modal-dialog textarea", By.CSS_SELECTOR)
        field.send_keys("yadissimo")
        path = "//button[contains(text(), 'Add Record')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        path = "//legend[contains(text(), 'Publication')]"
        legend = self.find(path, By.XPATH)
        self.assertIsNotNone(legend)
        legend.click()
        sleep(1)
        field = self.find("input[testid='PublisherID']", By.CSS_SELECTOR)
        field.send_keys("Test Company Name")
        sleep(1)
        field.send_keys(Keys.TAB)
        field = self.find("input[testid='SupplierID']", By.CSS_SELECTOR)
        field.send_keys("Test Company Name")
        sleep(1)
        field.send_keys(Keys.TAB)
        legend.click()
        sleep(1)
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        self.scroll_to_bottom()
        sleep(2)
        button.click()

    def test_05_handbell_ensemble_value(self):
        """Add a new value to the handbell ensemble lookup table."""

        button = self.find("//a[contains(text(), 'Lookup Tables')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        path = "a[href='/library/lookup/handbell-ensemble']"
        link = self.find(path, By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.assert_page_has("HandBell Ensembles")
        self.assert_page_has("Select a value to edit")
        button = self.find("//button[contains(text(), 'Create')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        self.assert_page_has("Editing Lookup Value")
        self.set_field_value("LookupValue", "Test Handbell Ensemble Value")
        self.set_field_value("SortPosition", "-1000")
        self.set_field_value("Comments", "Yada yada yada...")
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.set_field_value("title", "Test Item Title")
        button = self.find("//button[contains(text(), 'Search')]", By.XPATH)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        button = self.find("button[title='Edit item']", By.CSS_SELECTOR)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        path = "//legend[contains(text(), 'Musical Information')]"
        legend = self.find(path, By.XPATH)
        self.assertIsNotNone(legend)
        legend.click()
        sleep(1)
        field = self.find("HandbellEnsembleID", By.NAME)
        field.send_keys("Test Handbell Ensemble Value")
        sleep(1)
        xpath = "//option[contains(text(), 'Test Handbell Ensemble Value')]"
        option = self.find(xpath, By.XPATH)
        option.click()
        self.scroll_to_top()
        legend.click()
        sleep(1)
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(2)
        self.scroll_to_bottom()
        button.click()

    def test_06_key_value(self):
        """Add a new value to the key lookup table."""

        button = self.find("//a[contains(text(), 'Lookup Tables')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        path = "a[href='/library/lookup/key']"
        link = self.find(path, By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.assert_page_has("Keys")
        self.assert_page_has("Select a value to edit")
        button = self.find("//button[contains(text(), 'Create')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        self.assert_page_has("Editing Lookup Value")
        self.set_field_value("LookupValue", "Test Key Value")
        self.set_field_value("SortPosition", "-1000")
        self.set_field_value("Comments", "Yada yada yada...")
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.set_field_value("title", "Test Item Title")
        button = self.find("//button[contains(text(), 'Search')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        button = self.find("button[title='Edit item']", By.CSS_SELECTOR)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        path = "//legend[contains(text(), 'Musical Information')]"
        legend = self.find(path, By.XPATH)
        self.assertIsNotNone(legend)
        legend.click()
        sleep(1)
        field = self.find("KeyID", By.NAME)
        field.send_keys("Test Key Value")
        sleep(1)
        xpath = "//option[contains(text(), 'Test Key Value')]"
        option = self.find(xpath, By.XPATH)
        option.click()
        legend.click()
        sleep(1)
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(2)
        button.click()

    def test_07_keyword_value(self):
        """Add a new value to the keyword lookup table."""

        button = self.find("//a[contains(text(), 'Lookup Tables')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        path = "a[href='/library/lookup/keyword']"
        link = self.find(path, By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.assert_page_has("Keywords")
        self.assert_page_has("Select a value to edit")
        button = self.find("//button[contains(text(), 'Create')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        self.assert_page_has("Editing Lookup Value")
        self.set_field_value("LookupValue", "Test Keyword Value")
        self.set_field_value("Comments", "Yada yada yada...")
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.set_field_value("title", "Test Item Title")
        button = self.find("//button[contains(text(), 'Search')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        button = self.find("button[title='Edit item']", By.CSS_SELECTOR)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        path = "//legend[contains(text(), 'Identification')]"
        legend = self.find(path, By.XPATH)
        self.assertIsNotNone(legend)
        legend.click()
        sleep(1)
        field = self.find("input[testid='Keywords']", By.CSS_SELECTOR)
        field.send_keys("Test Keyword Value")
        sleep(1)
        field.send_keys(Keys.TAB)
        sleep(1)
        field.send_keys("justice")
        field.send_keys(Keys.TAB)
        legend.click()
        sleep(1)
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(2)
        button.click()

    def test_08_owner_value(self):
        """Add a new value to the owner lookup table."""

        button = self.find("//a[contains(text(), 'Lookup Tables')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        path = "a[href='/library/lookup/owner']"
        link = self.find(path, By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.assert_page_has("Owners")
        self.assert_page_has("Select a value to edit")
        button = self.find("//button[contains(text(), 'Create')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        self.assert_page_has("Editing Lookup Value")
        self.set_field_value("LookupValue", "Test Owner Value")
        self.set_field_value("SortPosition", "-1000")
        self.set_field_value("Comments", "Yada yada yada...")
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.set_field_value("title", "Test Item Title")
        button = self.find("//button[contains(text(), 'Search')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        button = self.find("button[title='Edit item']", By.CSS_SELECTOR)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        path = "//legend[contains(text(), 'Identification')]"
        legend = self.find(path, By.XPATH)
        self.assertIsNotNone(legend)
        legend.click()
        sleep(1)
        field = self.find("OwnerID", By.NAME)
        field.send_keys("Test Owner Value")
        sleep(1)
        xpath = "//option[contains(text(), 'Test Owner Value')]"
        option = self.find(xpath, By.XPATH)
        option.click()
        legend.click()
        sleep(1)
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(2)
        button.click()

    def test_09_person_value(self):
        """Add a new value to the person lookup table."""

        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.set_field_value("title", "Test Item Title")
        button = self.find("//button[contains(text(), 'Search')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        button = self.find("button[title='Edit item']", By.CSS_SELECTOR)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        path = "//button[contains(text(), 'Add Person')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(2)
        button.click()
        self.assert_page_has("Create New Person Record")
        self.set_field_value("LastName", "Person")
        self.set_field_value("FirstName", "Test")
        self.set_field_value("Dates", "1900-1999")
        field = self.find(".modal-dialog textarea", By.CSS_SELECTOR)
        field.send_keys("Long, productive career")
        path = "//button[contains(text(), 'Add Record')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        path = "//legend[contains(text(), 'Creation')]"
        legend = self.find(path, By.XPATH)
        self.assertIsNotNone(legend)
        legend.click()
        sleep(1)
        field = self.find("input[testid='ComposerID']", By.CSS_SELECTOR)
        field.send_keys("Person, Test")
        sleep(1)
        field.send_keys(Keys.TAB)
        field = self.find("input[testid='LyricistID']", By.CSS_SELECTOR)
        field.send_keys("Person, Test")
        sleep(1)
        field.send_keys(Keys.TAB)
        field = self.find("input[testid='ArrangerID']", By.CSS_SELECTOR)
        field.send_keys("Person, Test")
        sleep(1)
        field.send_keys(Keys.TAB)
        legend.click()
        sleep(1)
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(2)
        button.click()

    def test_10_season_value(self):
        """Add a new value to the season lookup table."""

        button = self.find("//a[contains(text(), 'Lookup Tables')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        path = "a[href='/library/lookup/season']"
        link = self.find(path, By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.assert_page_has("Seasons")
        self.assert_page_has("Select a value to edit")
        button = self.find("//button[contains(text(), 'Create')]", By.XPATH)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        self.assert_page_has("Editing Lookup Value")
        self.set_field_value("LookupValue", "Test Season Value")
        self.set_field_value("SortPosition", "-1000")
        self.set_field_value("Comments", "Yada yada yada...")
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.set_field_value("title", "Test Item Title")
        button = self.find("//button[contains(text(), 'Search')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        button = self.find("button[title='Edit item']", By.CSS_SELECTOR)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        path = "//legend[contains(text(), 'Identification')]"
        legend = self.find(path, By.XPATH)
        self.assertIsNotNone(legend)
        legend.click()
        sleep(1)
        field = self.find("SeasonID", By.NAME)
        field.send_keys("Test Season Value")
        sleep(1)
        xpath = "//option[contains(text(), 'Test Season Value')]"
        option = self.find(xpath, By.XPATH)
        option.click()
        self.scroll_to_top()
        legend.click()
        sleep(1)
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(2)
        button.click()

    def test_11_skill_value(self):
        """Add a new value to the skill/difficulty lookup table."""

        button = self.find("//a[contains(text(), 'Lookup Tables')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        path = "a[href='/library/lookup/skill']"
        link = self.find(path, By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.assert_page_has("Skills")
        self.assert_page_has("Select a value to edit")
        button = self.find("//button[contains(text(), 'Create')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        self.assert_page_has("Editing Lookup Value")
        self.set_field_value("LookupValue", "Test Skill Value")
        self.set_field_value("SortPosition", "-1000")
        self.set_field_value("Comments", "Yada yada yada...")
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.set_field_value("title", "Test Item Title")
        button = self.find("//button[contains(text(), 'Search')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        button = self.find("button[title='Edit item']", By.CSS_SELECTOR)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        path = "//legend[contains(text(), 'Identification')]"
        legend = self.find(path, By.XPATH)
        self.assertIsNotNone(legend)
        legend.click()
        sleep(1)
        field = self.find("SkillID", By.NAME)
        field.send_keys("Test Skill Value")
        sleep(1)
        xpath = "//option[contains(text(), 'Test Skill Value')]"
        option = self.find(xpath, By.XPATH)
        option.click()
        self.scroll_to_top()
        legend.click()
        sleep(1)
        path = "//legend[contains(text(), 'Musical Information')]"
        legend = self.find(path, By.XPATH)
        self.assertIsNotNone(legend)
        legend.click()
        sleep(1)
        voices = "Soprano", "Alto", "Tenor", "Bass"
        for voice in voices:
            field = self.find(f"{voice}ID", By.NAME)
            field.send_keys("Test Skill Value")
            sleep(1)
            xpath = ".//option[contains(text(), 'Test Skill Value')]"
            option = field.find_element(By.XPATH, xpath)
            option.click()
        legend.click()
        sleep(1)
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(2)
        button.click()

    def test_12_tag_value(self):
        """Add a new value to the tag lookup table."""

        button = self.find("//a[contains(text(), 'Lookup Tables')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        self.scroll_to_bottom()
        path = "a[href='/library/lookup/tag-group']"
        link = self.find(path, By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.assert_page_has("Tag Groups")
        self.assert_page_has("Select a value to edit")
        button = self.find("//button[contains(text(), 'Create')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        self.assert_page_has("Editing Lookup Value")
        self.set_field_value("TagGroupName", "Test Tag Group")
        self.set_field_value("Comments", "Yada yada yada...")
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        button = self.find("//a[contains(text(), 'Lookup Tables')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        self.scroll_to_bottom()
        path = "a[href='/library/lookup/tag']"
        link = self.find(path, By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.assert_page_has("Tags")
        self.assert_page_has("Select a value to edit")
        button = self.find("//button[contains(text(), 'Create')]", By.XPATH)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        self.assert_page_has("Editing Lookup Value")
        field = self.find("TagGroup", By.NAME)
        field.send_keys("Test Tag Group")
        sleep(1)
        xpath = "//option[contains(text(), 'Test Tag Group')]"
        option = self.find(xpath, By.XPATH)
        option.click()
        self.set_field_value("TagName", "Test Tag Name")
        self.set_field_value("Comments", "Yada yada yada...")
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.set_field_value("title", "Test Item Title")
        button = self.find("//button[contains(text(), 'Search')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        button = self.find("button[title='Edit item']", By.CSS_SELECTOR)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        path = "//legend[contains(text(), 'Identification')]"
        legend = self.find(path, By.XPATH)
        self.assertIsNotNone(legend)
        legend.click()
        sleep(1)
        field = self.find("input[testid='Tags']", By.CSS_SELECTOR)
        field.send_keys("Test Tag Group: Test Tag Name")
        sleep(1)
        field.send_keys(Keys.TAB)
        sleep(1)
        field.send_keys("ensemble: adult choir")
        field.send_keys(Keys.TAB)
        legend.click()
        sleep(1)
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(2)
        button.click()

    def test_13_performances(self):
        """Add performance blocks."""

        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.set_field_value("title", "Test Item Title")
        button = self.find("//button[contains(text(), 'Search')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        button = self.find("button[title='Edit item']", By.CSS_SELECTOR)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        legend_path = "//legend[contains(text(), 'Performances')]"
        self.scroll_to_bottom()
        self.find(legend_path, By.XPATH).click()
        sleep(1)
        add_performance_path = "//button[contains(text(), 'Add Performance')]"
        self.find(add_performance_path, By.XPATH).click()
        self.set_field_value("Performances__0__PerformanceDate", "12252020")
        self.set_field_value("Performances__0__Comments", "Awesome!")
        button = self.find(add_performance_path, By.XPATH)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        self.scroll_to_bottom()
        button.click()
        self.set_field_value("Performances__1__PerformanceDate", "12252022")
        self.set_field_value("Performances__1__Comments", "Super!")
        button = self.find(add_performance_path, By.XPATH)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        self.scroll_to_bottom()
        button.click()
        self.set_field_value("Performances__2__PerformanceDate", "12252024")
        self.set_field_value("Performances__2__Comments", "Inspiring!")
        selector = "legend[href='#nested-Performances-1'] button.trash-button"
        self.find(selector, By.CSS_SELECTOR).click()
        self.scroll_to_bottom()
        self.find(".modal-footer .btn-danger", By.CSS_SELECTOR).click()
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(2)
        button.click()

    def test_14_inventories(self):
        """Add inventory blocks."""

        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        sleep(1)
        self.set_field_value("title", "Test Item Title")
        button = self.find("//button[contains(text(), 'Search')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        sleep(1)
        button = self.find("button[title='Edit item']", By.CSS_SELECTOR)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        sleep(1)
        legend_path = "//legend[contains(text(), 'Inventories')]"
        self.scroll_to_bottom()
        self.find(legend_path, By.XPATH).click()
        sleep(1)
        add_inventory_path = "//button[contains(text(), 'Add Inventory')]"
        self.find(add_inventory_path, By.XPATH).click()
        sleep(1)
        self.set_field_value("Inventories__0__InStock", "50")
        self.set_field_value("Inventories__0__InStockDate", "01012020")
        self.set_field_value("Inventories__0__StorageLocation", "Shelf 1")
        self.set_field_value("Inventories__0__LatestPrice", "2.50")
        self.set_field_value("Inventories__0__AcquireCondition", "New")
        self.set_field_value("Inventories__0__Comments", "yada yada yada")
        path = "legend[href='#nested-Inventories-0']"
        self.find(path, By.CSS_SELECTOR).click()
        sleep(1)
        button = self.find(add_inventory_path, By.XPATH)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        button.click()
        sleep(1)
        self.set_field_value("Inventories__1__InStock", "50")
        self.set_field_value("Inventories__1__InStockDate", "01012022")
        self.set_field_value("Inventories__1__StorageLocation", "Floor")
        self.set_field_value("Inventories__1__LatestPrice", "2.50")
        self.set_field_value("Inventories__1__AcquireCondition", "OK")
        self.set_field_value("Inventories__1__Comments", "yada yada yada")
        path = "legend[href='#nested-Inventories-1']"
        self.find(path, By.CSS_SELECTOR).click()
        sleep(1)
        button = self.find(add_inventory_path, By.XPATH)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(1)
        button.click()
        sleep(1)
        self.set_field_value("Inventories__2__InStock", "60")
        self.set_field_value("Inventories__2__InStockDate", "01012024")
        self.set_field_value("Inventories__2__StorageLocation", "Shelf 2")
        self.set_field_value("Inventories__2__LatestPrice", "2.75")
        self.set_field_value("Inventories__2__AcquireCondition", "Shabby")
        self.set_field_value("Inventories__2__Comments", "yada yada yada")
        path = "legend[href='#nested-Inventories-2']"
        self.find(path, By.CSS_SELECTOR).click()
        sleep(1)
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(1)
        button.click()

    def test_15_parts(self):
        """Add part blocks."""

        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.set_field_value("title", "Test Item Title")
        button = self.find("//button[contains(text(), 'Search')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        sleep(1)
        button = self.find("button[title='Edit item']", By.CSS_SELECTOR)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        sleep(1)
        legend_path = "//legend[contains(text(), 'Parts')]"
        self.scroll_to_bottom()
        self.find(legend_path, By.XPATH).click()
        sleep(1)
        add_part_path = "//button[contains(text(), 'Add Part')]"
        self.find(add_part_path, By.XPATH).click()
        sleep(1)
        self.set_field_value("Parts__0__PartName", "oboe")
        self.set_field_value("Parts__0__InventoryDate", "01012020")
        self.set_field_value("Parts__0__OnHand", "1")
        self.set_field_value("Parts__0__Needed", "2")
        path = "legend[href='#nested-Parts-0']"
        self.find(path, By.CSS_SELECTOR).click()
        sleep(1)
        self.find(add_part_path, By.XPATH).click()
        sleep(1)
        self.set_field_value("Parts__1__PartName", "clarinet")
        self.set_field_value("Parts__1__InventoryDate", "01012020")
        self.set_field_value("Parts__1__OnHand", "1")
        path = "legend[href='#nested-Parts-1']"
        self.find(path, By.CSS_SELECTOR).click()
        sleep(2)
        button = self.find(add_part_path, By.XPATH)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        button.click()
        sleep(1)
        self.set_field_value("Parts__2__PartName", "bassoon")
        self.set_field_value("Parts__2__InventoryDate", "01012020")
        self.set_field_value("Parts__2__OnHand", "1")
        path = "legend[href='#nested-Parts-2']"
        self.find(path, By.CSS_SELECTOR).click()
        sleep(1)
        button = self.find(add_part_path, By.XPATH)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        button.click()
        sleep(1)
        self.set_field_value("Parts__3__PartName", "French horn")
        self.set_field_value("Parts__3__InventoryDate", "01012020")
        self.set_field_value("Parts__3__OnHand", "1")
        path = "legend[href='#nested-Parts-3']"
        self.find(path, By.CSS_SELECTOR).click()
        sleep(1)
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(1)
        button.click()

    def test_16_loans(self):
        """Add loan blocks."""

        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.set_field_value("title", "Test Item Title")
        button = self.find("//button[contains(text(), 'Search')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        button = self.find("button[title='Edit item']", By.CSS_SELECTOR)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        legend_path = "//legend[contains(text(), 'Loans')]"
        self.scroll_to_bottom()
        self.find(legend_path, By.XPATH).click()
        sleep(1)
        add_loan_path = "//button[contains(text(), 'Add Loan')]"
        self.find(add_loan_path, By.XPATH).click()
        self.set_field_value("Loans__0__LoanRecipient", "Klem Kadiddlehopper")
        self.set_field_value("Loans__0__LoanDate", "01012020")
        self.set_field_value("Loans__0__LoanReturned", "01012022")
        self.set_field_value("Loans__0__Comments", "yada yada yada")
        path = "legend[href='#nested-Loans-0']"
        self.find(path, By.CSS_SELECTOR).click()
        sleep(1)
        path = "//button[contains(text(), 'Save')]"
        button = self.find(path, By.XPATH)
        self.assertIsNotNone(button)
        self.driver.execute_script("arguments[0].scrollIntoView(true)", button)
        sleep(2)
        button.click()

    def test_17_title_report(self):
        """Filter catalog items by title and generate a report."""

        self.assert_page_has("Browse Catalog")
        legend_path = "//legend[contains(text(), 'Options')]"
        legend = self.find(legend_path, By.XPATH)
        legend.click()
        sleep(1)
        self.set_field_value("report-title", "Test Report")
        path = "//div[@class='rbt-token-label' and text()='Title']"
        target = self.find(path, By.XPATH)
        self.assertIsNotNone(target)
        self.assertTrue(target.is_displayed())
        path = "//div[@class='rbt-token-label' and text()='Item ID']"
        draggable = self.find(path, By.XPATH)
        self.assertIsNotNone(draggable)
        self.assertTrue(draggable.is_displayed())
        action_chain = ActionChains(self.driver)
        # Doesn't work; do the drag and drop manually.
        # action_chain.drag_and_drop(draggable, target).perform()
        action_chain.click_and_hold(draggable)
        action_chain.move_to_element(target)
        action_chain.release(target)
        action_chain.perform()
        field = self.find("input[testid='report-columns']", By.CSS_SELECTOR)
        field.send_keys("other")
        field.send_keys(Keys.TAB)
        legend.click()
        legend_path = "//legend[contains(text(), 'Identification')]"
        legend = self.find(legend_path, By.XPATH)
        self.scroll_to_bottom()
        legend.click()
        sleep(1)
        self.set_field_value("title", "Test Item Title")
        legend.click()
        sleep(1)
        original = self.driver.current_window_handle
        self.run_report()
        report = self.driver.current_window_handle
        self.assert_page_has("Test Report")
        self.save_pdf("report.pdf")
        self.find("td a", By.CSS_SELECTOR).click()
        handles = self.driver.window_handles
        for handle in handles:
            if handle not in (original, report):
                self.driver.switch_to.window(handle)
                break
        sleep(1)
        self.save_pdf("link-from-report.pdf")

    def test_18_composer_report(self):
        """Filter catalog items by composer and generate a report."""

        self.assert_page_has("Browse Catalog")
        legend_path = "//legend[contains(text(), 'Identification')]"
        legend = self.find(legend_path, By.XPATH)
        self.scroll_to_bottom()
        legend.click()
        sleep(1)
        self.set_field_value("creator", "Person, Test")
        legend.click()
        sleep(1)
        self.run_report("Person, Test")

    def test_19_arrangement_report(self):
        """Filter catalog items by arrangement and generate a report."""

        self.assert_page_has("Browse Catalog")
        legend_path = "//legend[contains(text(), 'Identification')]"
        legend = self.find(legend_path, By.XPATH)
        self.scroll_to_bottom()
        legend.click()
        sleep(1)
        field = self.find("input[testid='arrangement']", By.CSS_SELECTOR)
        field.send_keys("Test Arrangement Value")
        sleep(1)
        field.send_keys(Keys.TAB)
        legend.click()
        sleep(1)
        self.run_report()

    def test_20_keyword_report(self):
        """Filter catalog items by keyword and generate a report."""

        self.assert_page_has("Browse Catalog")
        legend_path = "//legend[contains(text(), 'Classification')]"
        legend = self.find(legend_path, By.XPATH)
        self.scroll_to_bottom()
        legend.click()
        field = self.find("input[testid='keyword']", By.CSS_SELECTOR)
        field.send_keys("Test Keyword Value")
        sleep(1)
        field.send_keys(Keys.TAB)
        legend.click()
        sleep(1)
        self.run_report()

    def test_21_occasion_report(self):
        """Filter catalog items by occasion and generate a report."""

        self.assert_page_has("Browse Catalog")
        legend_path = "//legend[contains(text(), 'Classification')]"
        legend = self.find(legend_path, By.XPATH)
        self.scroll_to_bottom()
        legend.click()
        sleep(1)
        field = self.find("input[testid='season']", By.CSS_SELECTOR)
        field.send_keys("Test Season Value")
        sleep(1)
        field.send_keys(Keys.TAB)
        legend.click()
        sleep(1)
        self.run_report()

    def test_22_tag_report(self):
        """Filter catalog items by tag and generate a report."""

        self.assert_page_has("Browse Catalog")
        legend_path = "//legend[contains(text(), 'Classification')]"
        legend = self.find(legend_path, By.XPATH)
        self.scroll_to_bottom()
        legend.click()
        sleep(1)
        field = self.find("input[testid='tag']", By.CSS_SELECTOR)
        field.send_keys("Test Tag")
        sleep(1)
        field.send_keys(Keys.TAB)
        legend.click()
        sleep(1)
        self.run_report()

    def test_23_owner_report(self):
        """Filter catalog items by owner and generate a report."""

        self.assert_page_has("Browse Catalog")
        legend_path = "//legend[contains(text(), 'Activity')]"
        legend = self.find(legend_path, By.XPATH)
        self.scroll_to_bottom()
        legend.click()
        sleep(1)
        field = self.find("input[testid='owner']", By.CSS_SELECTOR)
        field.send_keys("Test Owner Value")
        sleep(1)
        field.send_keys(Keys.TAB)
        legend.click()
        sleep(1)
        self.run_report()

    def test_24_excel_report(self):
        """Test the Excel version of a report."""

        self.assert_page_has("Browse Catalog")
        self.scroll_to_bottom()
        legend_path = "//legend[contains(text(), 'Options')]"
        legend = self.find(legend_path, By.XPATH)
        legend.click()
        sleep(1)
        self.find("report-format-excel", By.ID).click()
        legend.click()
        sleep(1)
        self.run_report(None)
        page = self.get_page_source()
        pattern = re_compile(r"Report Music Library Report \d{8}-(\d+)\.xlsx")
        match = pattern.search(page)
        self.assertIsNotNone(match)
        request_id = match.group(1)
        credentials = {
            "username": self.USERNAME,
            "password": self.PASSWORD,
        }
        session = Session()
        api = f"{self.url}/api"
        response = session.post(f"{api}/session", json=credentials).json()
        self.assertEqual(response["status"], "success")
        url = f"{api}/report/{request_id}"
        self.logger.info("fetching %s", url)
        response = session.get(url)
        book = load_workbook(BytesIO(response.content))
        sheet = book.active
        found = False
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == "Test Item Title":
                    found = True
                    break
            if found:
                break
        self.assertTrue(found)

    def test_25_print(self):
        """Check the full record for the item we created."""

        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.scroll_to_bottom()
        self.set_field_value("title", "Test Item Title")
        button = self.find("//button[contains(text(), 'Search')]", By.XPATH)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        self.scroll_to_bottom()
        selector = "button[title='Display printable version of item']"
        button = self.find(selector, By.CSS_SELECTOR)
        self.assertIsNotNone(button)
        button.click()
        sleep(1)
        handles = self.driver.window_handles
        current = self.driver.current_window_handle
        self.logger.info("handles=%s current=%s", handles, current)
        for handle in handles:
            if handle != current:
                self.driver.switch_to.window(handle)
                break
        self.save_pdf("printable-item.pdf")
        self.assert_page_has("Test Item Title [item ID")
        self.assert_page_has("Test Item Other Title")
        self.assert_page_has("Test Skill Value")
        self.assert_page_has("Test Owner Value")
        self.assert_page_has("Adult Choir; Test Tag Name")
        self.assert_page_has("Test Item Comments")
        self.assert_page_has("Person, Test (1900-1999)")
        self.assert_page_has("Test Company Name")
        self.assert_page_has("2025 All Rights Reserved")
        self.assert_page_has("Test Item Stock Number")
        self.assert_page_has("Test Key Value")
        self.assert_page_has("Test Accompaniment Value")
        self.assert_page_has("Test Arrangement Value")
        self.assert_page_has("5.25")
        self.assert_page_has("3/4")
        self.assert_page_has("Test Handbell Ensemble Value")
        self.assert_page_has("Test Item Program Notes")
        self.assert_page_has("2020-12-25")
        self.assert_page_has("Awesome!")
        self.assert_page_has("2024-12-25")
        self.assert_page_has("Inspiring!")
        self.assert_page_not_has("Super!")
        self.assert_page_has("50")
        self.assert_page_has("2020-01-01")
        self.assert_page_has("Shelf 1")
        self.assert_page_has("2.50")
        self.assert_page_has("New")
        self.assert_page_has("yada yada yada")
        self.assert_page_has("2022-01-01")
        self.assert_page_has("Floor")
        self.assert_page_has("OK")
        self.assert_page_has("2024-01-01")
        self.assert_page_has("Shelf 2")
        self.assert_page_has("2.75")
        self.assert_page_has("Shabby")
        self.assert_page_has("oboe")
        self.assert_page_has("clarinet")
        self.assert_page_has("bassoon")
        self.assert_page_has("French horn")
        self.assert_page_has("Klem Kadiddlehopper")

    def test_26_cleanup(self):
        """Remove our test data."""

        link = self.find("a[href='/library/edit']", By.CSS_SELECTOR)
        self.assertIsNotNone(link)
        link.click()
        self.set_field_value("title", "Test Item Title")
        button = self.find("//button[contains(text(), 'Search')]", By.XPATH)
        self.assertIsNotNone(button)
        button.click()
        button = self.find("button[title='Delete this item']", By.CSS_SELECTOR)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        button = self.find("//button[contains(text(), 'Delete')]", By.XPATH)
        self.assertIsNotNone(button)
        self.scroll_to_bottom()
        button.click()
        self.assert_page_has("Successfully deleted Test Item Title.")
        self.assert_page_has("No matching records found.")
        for name, label, value in self.LOOKUPS:
            path = "//a[contains(text(), 'Lookup Tables')]"
            button = self.find(path, By.XPATH)
            self.assertIsNotNone(button)
            button.click()
            self.scroll_to_bottom()
            path = f"a[href='/library/lookup/{name}']"
            link = self.find(path, By.CSS_SELECTOR)
            self.assertIsNotNone(link)
            link.click()
            self.assert_page_has(label)
            self.assert_page_has("Select a value to edit")
            field = self.find("lookup-values", By.NAME)
            field.send_keys(value)
            xpath = f"//option[contains(text(), '{value}')]"
            option = self.find(xpath, By.XPATH)
            option.click()
            button = self.find("//button[contains(text(), 'Edit')]", By.XPATH)
            self.assertIsNotNone(button)
            button.click()
            selector = "button[title='Delete this value']"
            button = self.find(selector, By.CSS_SELECTOR)
            self.assertIsNotNone(button)
            script = "arguments[0].scrollIntoView(true)"
            self.driver.execute_script(script, button)
            sleep(2)
            self.scroll_to_bottom()
            button.click()
            sleep(1)
            self.assert_page_has("Are you sure you want to delete this value?")
            selector = ".modal-footer button.btn-danger"
            button = self.find(selector, By.CSS_SELECTOR)
            self.assertIsNotNone(button)
            button.click()
            sleep(1)


if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument("--username", "-u", required=True)
    parser.add_argument("--password", "-p", required=True)
    parser.add_argument("--base", "-b")
    parser.add_argument("--verbose", "-v", action="store_true")
    parser.add_argument("--tests", "-t", nargs="*")
    opts = parser.parse_args()
    Tests.USERNAME = opts.username
    Tests.PASSWORD = opts.password
    if opts.base:
        Tests.BASE = opts.base
    if opts.verbose:
        Tests.VERBOSE = True
    Tests.LOGGER.info("-" * 40)
    Tests.LOGGER.info("Tests started using %s", Tests.BASE)
    new_args = ["-v"] if opts.verbose else []
    if opts.tests:
        new_args = opts.tests
    argv[1:] = new_args
    del opts
    del parser
    main()
    elapsed = datetime.now() - Tests.STARTED
    Tests.LOGGER.info("Tests completed in %s", elapsed)
